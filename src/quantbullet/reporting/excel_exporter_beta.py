import os
import pandas as pd
from pyparsing import col
from .columns import ColumnFormat
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

class TableBlock:
    def __init__(
        self,
        df,
        start_row=None,
        start_col=None,
        column_formats=None,
        wrap_header=False,
        include_header=True,
        row_offset=1,
        col_offset=0
    ):
        self.df = df.copy()
        self.start_row = start_row
        self.start_col = start_col
        self.column_formats = column_formats or {col: ColumnFormat() for col in df.columns}
        self.wrap_header = wrap_header
        self.include_header = include_header
        self.row_offset = row_offset
        self.col_offset = col_offset

class ExcelExporterBeta:
    def __init__(self, filename):
        self.filename = filename
        self._overwrite = False
        self.default_alignment = Alignment(horizontal='right')
        self._sheets = {}

    @property
    def overwrite(self):
        return self._overwrite

    @overwrite.setter
    def overwrite(self, value):
        self.set_overwrite(value)

    def set_overwrite(self, value: bool):
        """Set whether to overwrite the existing Excel file or not.
        
        If you want to reuse the same file, just write to different sheets, then set this to False.
        """
        self._overwrite = bool(value)
        return self
    
    def add_table(self, sheet_name, df, start_row=1, start_col=1, column_formats=None, wrap_header=False, include_header=True, row_offset=1, col_offset=0):
        if self._is_duplicate_columns(df):
            raise ValueError("DataFrame contains duplicate columns. Please rename them before exporting to Excel.")

        # Initialize or complete column_formats
        if column_formats is None:
            column_formats = {col: ColumnFormat() for col in df.columns}
        else:
            # Validate and fill missing columns
            for col in column_formats:
                if col not in df.columns:
                    raise ValueError(f"Column '{col}' in column_formats does not exist in DataFrame.")
            for col in df.columns:
                if col not in column_formats:
                    column_formats[col] = ColumnFormat()

        block = TableBlock(
            df,
            start_row=start_row,
            start_col=start_col,
            column_formats=column_formats,
            wrap_header=wrap_header,
            include_header=include_header,
            row_offset=row_offset,
            col_offset=col_offset
        )

        if sheet_name not in self._sheets:
            self._sheets[sheet_name] = []
        self._sheets[sheet_name].append(block)
        return self

    @staticmethod
    def _is_duplicate_columns(df):
        """Check if the DataFrame has duplicate columns."""
        return df.columns.duplicated().any()
    
    def _build_number_format(self, decimals, fmt: ColumnFormat):
        base = '0' + ('.' + '0' * decimals if decimals > 0 else '')
        
        if fmt is None:
            return base
        
        if fmt.comma:
            base = "#,##" + base
            
        if fmt.percent:
            base += "%"
                
        # Add negative in parentheses
        if fmt.parens_for_negative:
            # Excel format: positive;negative;zero
            base = f"{base};({base});{base}"
        
        return base

    def _get_col_format_strings(self, df, col_formats):
        """Get the column format strings for the DataFrame."""
        formats = {}
        for col in df.columns:
            fmt = col_formats.get(col, None)
            if pd.api.types.is_numeric_dtype(df[col]):
                # fmt is a type of ColumnFormat and should have decimals as an attribute and default values
                decimals = fmt.decimals if fmt and fmt.decimals is not None else None
                pattern = self._build_number_format(decimals, fmt)
                formats[col] = pattern
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                formats[col] = fmt.date_format if fmt and fmt.date_format is not None else None
        return formats

    def _apply_formatting(self, worksheet, df, format_strings, column_formats, start_row=1, start_col=1, wrap_header=False, include_header=True):
        data_start_row = start_row + 1 if include_header else start_row
        for j, col in enumerate(df.columns):
            col_index = start_col + j
            col_letter = get_column_letter(col_index)
            
            column_format = column_formats.get(col)

            # Set width
            if column_format and column_format.width is not None:
                # HACK It seems there is a gap in column width we set and the actual width exported. and this gap is known to be 0.7.
                width = column_format.width + 0.7
            elif column_format:
                width = column_format.estimate_display_width(df[col], col, column_format.decimals) + 2
            else:
                width = max(df[col].astype(str).map(len).max(), len(col)) + 2  # fallback
            worksheet.column_dimensions[col_letter].width = width

            # Format data cells
            for i in range(len(df)):
                cell = worksheet.cell(row=data_start_row + i, column=col_index)
                cell.alignment = self.default_alignment
                if format_strings.get(col):
                    cell.number_format = format_strings[col]

            # Wrap header text
            if wrap_header and include_header:
                header_cell = worksheet.cell(row=start_row, column=col_index)
                header_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                worksheet.row_dimensions[start_row].height = 30
                
            # Conditional formatting
            if column_format and column_format.color_scale:
                rule = column_format.build_conditional_formatting_rule()
                if rule:
                    cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_start_row + len(df)-1}"
                    worksheet.conditional_formatting.add(cell_range, rule)
                    
            # Formula Fill
            if column_format and column_format.formula_template:
                for i in range(len(df)):
                    formula = column_format.formula_template.format(row=data_start_row + i)
                    worksheet.cell(row=data_start_row + i, column=col_index).value = formula
            
    def _apply_column_transforms(self, df, col_formats):
        """Apply transformations to the DataFrame columns based on the provided formats."""
        for col, fmt in col_formats.items():
            if col in df.columns:
                df[col] = fmt.apply_transform(df[col])
        return df

    def save(self, drop_existing_sheets: bool = True):
        """Save the Excel file with optional sheet replacement for first table only."""
        mode = 'w' if self.overwrite else ('a' if os.path.exists(self.filename) else 'w')

        # Main writer to be used for all overlay-style writes
        writer = pd.ExcelWriter(self.filename, engine="openpyxl", mode=mode)

        try:
            for sheet_name, blocks in self._sheets.items():
                current_row = 1
                current_col = 1

                for i, block in enumerate(blocks):
                    df = self._apply_column_transforms(block.df, block.column_formats)
                    format_strings = self._get_col_format_strings(df, block.column_formats)

                    # Determine actual position
                    actual_start_row = block.start_row if block.start_row is not None else current_row + block.row_offset
                    actual_start_col = block.start_col if block.start_col is not None else current_col + block.col_offset

                    # Decide writer
                    if i == 0 and drop_existing_sheets and mode == 'a':
                        # First block on this sheet and we're dropping — use temp writer with replace
                        temp_writer = pd.ExcelWriter(
                            self.filename,
                            engine="openpyxl",
                            mode="a",
                            if_sheet_exists="replace"
                        )
                        df.to_excel(
                            temp_writer,
                            sheet_name=sheet_name,
                            index=False,
                            header=block.include_header,
                            startrow=actual_start_row - 1,
                            startcol=actual_start_col - 1
                        )
                        temp_writer.close()

                        # Refresh the main writer to reflect the change
                        writer = pd.ExcelWriter(self.filename, engine="openpyxl", mode="a", if_sheet_exists='overlay')
                    else:
                        # Normal case: write with overlay
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False,
                            header=block.include_header,
                            startrow=actual_start_row - 1,
                            startcol=actual_start_col - 1,
                        )

                    # Retrieve worksheet
                    worksheet = writer.sheets[sheet_name]

                    self._apply_formatting(
                        worksheet,
                        df,
                        format_strings,
                        block.column_formats,
                        start_row=actual_start_row,
                        start_col=actual_start_col,
                        wrap_header=block.wrap_header,
                        include_header=block.include_header
                    )

                    # Track position
                    rows_written = len(df) + (1 if block.include_header else 0)
                    cols_written = len(df.columns)
                    current_row = actual_start_row + rows_written - 1
                    current_col = actual_start_col

        except Exception as e:
            print(f"Error while exporting Excel file: {e}")
            raise
        finally:
            writer.close()
            self._sheets.clear()