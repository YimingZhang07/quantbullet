import os
import pandas as pd
from .columns import ColumnFormat
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from .columns import ColumnSchema

class ExcelExporter:
    def __init__(self, filename):
        self.filename = filename
        self._overwrite = False
        self.default_alignment = Alignment(horizontal='right')
        self._sheets = []

    @property
    def overwrite(self):
        return self._overwrite

    @overwrite.setter
    def overwrite(self, value):
        self.set_overwrite(value)

    def set_overwrite(self, value: bool):
        """Set whether to overwrite the existing Excel file or not.
        
        If you want to reuse the same file, just write to different sheets, then set this to False.
        """
        self._overwrite = bool(value)
        return self

    def add_sheet(self, sheet_name, df, schema: ColumnSchema = None, start_col=1, start_row=1, wrap_header=False, include_header=True):
        # Check for duplicate columns
        if self._is_duplicate_columns(df):
            raise ValueError("DataFrame contains duplicate columns. Please rename them before exporting to Excel.")
        
        df = df.copy()
        if schema:
            df.rename(columns=schema.rename_map, inplace=True)
            df = df[list(schema.col_config.keys())]
            column_formats = schema.col_config
        else:
            column_formats = None
        
        if column_formats is None:
            column_formats = {col: ColumnFormat() for col in df.columns}
        for col in column_formats.keys():
            if col not in df.columns:
                raise ValueError(f"Column '{col}' in column_formats does not exist in the DataFrame columns.")
        for col in df.columns:
            if col not in column_formats:
                column_formats[col] = ColumnFormat()

        self._sheets.append({
            "df": df.copy(),
            "sheet_name": sheet_name,
            "column_formats": column_formats or {},
            "wrap_header": wrap_header,
            "include_header": include_header,
            "start_col": start_col,
            "start_row": start_row
        })
        return self
    
    @staticmethod
    def _is_duplicate_columns(df):
        """Check if the DataFrame has duplicate columns."""
        return df.columns.duplicated().any()
    
    def _build_number_format(self, decimals, fmt: ColumnFormat):
        base = '0' + ('.' + '0' * decimals if decimals > 0 else '')
        
        if fmt is None:
            return base
        
        if fmt.comma:
            base = "#,##" + base
            
        if fmt.percent:
            base += "%"
                
        # Add negative in parentheses
        if fmt.parens_for_negative:
            # Excel format: positive;negative;zero
            base = f"{base};({base});{base}"
        
        return base

    def _get_col_format_strings(self, df, col_formats):
        """Get the column format strings for the DataFrame."""
        formats = {}
        for col in df.columns:
            fmt = col_formats.get(col, None)
            if pd.api.types.is_numeric_dtype(df[col]):
                # fmt is a type of ColumnFormat and should have decimals as an attribute and default values
                decimals = fmt.decimals if fmt and fmt.decimals is not None else None
                pattern = self._build_number_format(decimals, fmt)
                formats[col] = pattern
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                formats[col] = fmt.date_format if fmt and fmt.date_format is not None else None
        return formats

    def _apply_formatting(self, worksheet, df, format_strings, start_row=1, start_col=1, wrap_header=False, include_header=True):
        data_start_row = start_row + 1 if include_header else start_row
        for j, col in enumerate(df.columns):
            col_index = start_col + j
            col_letter = get_column_letter(col_index)
            
            sheet_config = next((s for s in self._sheets if s["df"] is df), None)
            column_format = sheet_config["column_formats"].get(col) if sheet_config else None

            # Set width
            if column_format and column_format.width is not None:
                # HACK It seems there is a gap in column width we set and the actual width exported. and this gap is known to be 0.7.
                width = column_format.width + 0.7
            elif column_format:
                width = column_format.estimate_display_width(df[col], col, column_format.decimals) + 2
            else:
                width = max(df[col].astype(str).map(len).max(), len(col)) + 2  # fallback
            worksheet.column_dimensions[col_letter].width = width

            # Format data cells
            for i in range(len(df)):
                cell = worksheet.cell(row=data_start_row + i, column=col_index)
                cell.alignment = self.default_alignment
                if format_strings.get(col):
                    cell.number_format = format_strings[col]

            # Wrap header text
            if wrap_header and include_header:
                header_cell = worksheet.cell(row=start_row, column=col_index)
                header_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                worksheet.row_dimensions[start_row].height = 30
                
            # Conditional formatting
            if column_format and column_format.color_scale:
                rule = column_format.build_conditional_formatting_rule()
                if rule:
                    cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_start_row + len(df)-1}"
                    worksheet.conditional_formatting.add(cell_range, rule)
                    
            # Formula Fill
            if column_format and column_format.formula_template:
                for i in range(len(df)):
                    formula = column_format.formula_template.format(row=data_start_row + i)
                    worksheet.cell(row=data_start_row + i, column=col_index).value = formula
            
    def _apply_column_transforms(self, df, col_formats):
        """Apply transformations to the DataFrame columns based on the provided formats."""
        for col, fmt in col_formats.items():
            if col in df.columns:
                df[col] = fmt.apply_transform(df[col])
        return df

    def save(self):
        """"Save the DataFrames to an Excel file."""
        # HACK the below condition does not make sense, cause if the file exists, it will be overwritten anyway.
        mode = 'w' if self.overwrite else ('a' if os.path.exists(self.filename) else 'w')
        writer_args = {"engine": "openpyxl", "mode": mode}
        if mode == 'a':
            writer_args["if_sheet_exists"] = "replace"
            
        writer = pd.ExcelWriter(self.filename, **writer_args)
        
        try:
            for sheet in self._sheets:
                df = sheet["df"]
                name = sheet["sheet_name"]
                col_formats = sheet["column_formats"]
                include_header = sheet["include_header"]
                start_col = sheet["start_col"]
                start_row = sheet["start_row"]
                # for any transformations, we need to apply them before formatting
                df = self._apply_column_transforms(df, col_formats)
                format_strings = self._get_col_format_strings(df, col_formats)
                df.to_excel(writer, sheet_name=name, index=False, header=include_header, startrow=start_row-1, startcol=start_col-1)
                worksheet = writer.sheets[name]
                self._apply_formatting(worksheet, df, format_strings, start_row, start_col, sheet["wrap_header"], sheet["include_header"])
        except Exception as e:
            print( f"Error while exporting Excel file: {e}" )
            raise
        finally:
            writer.close()
            self._sheets.clear()